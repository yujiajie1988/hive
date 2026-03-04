"""Tests for excel_tool - Read and manipulate Excel files (.xlsx, .xlsm)."""

import importlib.util
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

import pytest
from fastmcp import FastMCP

openpyxl_available = importlib.util.find_spec("openpyxl") is not None

# Skip all tests if openpyxl is not installed
pytestmark = pytest.mark.skipif(not openpyxl_available, reason="openpyxl not installed")

if openpyxl_available:
    from openpyxl import Workbook

    from aden_tools.tools.excel_tool.excel_tool import register_tools

# Test IDs for sandbox
TEST_WORKSPACE_ID = "test-workspace"
TEST_AGENT_ID = "test-agent"
TEST_SESSION_ID = "test-session"


@pytest.fixture
def excel_tools(mcp: FastMCP, tmp_path: Path):
    """Register all Excel tools and return them as a dict."""
    with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
        register_tools(mcp)
        yield {
            "excel_read": mcp._tool_manager._tools["excel_read"].fn,
            "excel_write": mcp._tool_manager._tools["excel_write"].fn,
            "excel_append": mcp._tool_manager._tools["excel_append"].fn,
            "excel_info": mcp._tool_manager._tools["excel_info"].fn,
            "excel_sheet_list": mcp._tool_manager._tools["excel_sheet_list"].fn,
            "excel_sql": mcp._tool_manager._tools["excel_sql"].fn,
            "excel_search": mcp._tool_manager._tools["excel_search"].fn,
        }


@pytest.fixture
def excel_read_fn(excel_tools):
    """Return excel_read function for backward compatibility."""
    return excel_tools["excel_read"]


@pytest.fixture
def session_dir(tmp_path: Path) -> Path:
    """Create and return the session directory within the sandbox."""
    session_path = tmp_path / TEST_WORKSPACE_ID / TEST_AGENT_ID / TEST_SESSION_ID
    session_path.mkdir(parents=True, exist_ok=True)
    return session_path


@pytest.fixture
def basic_xlsx(session_dir: Path) -> Path:
    """Create a basic Excel file for testing."""
    xlsx_file = session_dir / "basic.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Header row
    ws.append(["name", "age", "city"])
    # Data rows
    ws.append(["Alice", 30, "NYC"])
    ws.append(["Bob", 25, "LA"])
    ws.append(["Charlie", 35, "Chicago"])
    wb.save(xlsx_file)
    wb.close()
    return xlsx_file


@pytest.fixture
def multi_sheet_xlsx(session_dir: Path) -> Path:
    """Create an Excel file with multiple sheets."""
    xlsx_file = session_dir / "multi_sheet.xlsx"
    wb = Workbook()

    # First sheet (active)
    ws1 = wb.active
    ws1.title = "Employees"
    ws1.append(["id", "name", "department"])
    ws1.append([1, "Alice", "Engineering"])
    ws1.append([2, "Bob", "Marketing"])

    # Second sheet
    ws2 = wb.create_sheet("Products")
    ws2.append(["id", "name", "price"])
    ws2.append([1, "Widget", 99.99])
    ws2.append([2, "Gadget", 149.99])

    # Third sheet
    ws3 = wb.create_sheet("Summary")
    ws3.append(["metric", "value"])
    ws3.append(["total_employees", 2])
    ws3.append(["total_products", 2])

    wb.save(xlsx_file)
    wb.close()
    return xlsx_file


@pytest.fixture
def large_xlsx(session_dir: Path) -> Path:
    """Create a larger Excel file for pagination testing."""
    xlsx_file = session_dir / "large.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["id", "value"])
    for i in range(100):
        ws.append([i, i * 10])
    wb.save(xlsx_file)
    wb.close()
    return xlsx_file


@pytest.fixture
def empty_xlsx(session_dir: Path) -> Path:
    """Create an empty Excel file."""
    xlsx_file = session_dir / "empty.xlsx"
    wb = Workbook()
    wb.save(xlsx_file)
    wb.close()
    return xlsx_file


@pytest.fixture
def headers_only_xlsx(session_dir: Path) -> Path:
    """Create an Excel file with only headers."""
    xlsx_file = session_dir / "headers_only.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["name", "age", "city"])
    wb.save(xlsx_file)
    wb.close()
    return xlsx_file


@pytest.fixture
def xlsx_with_dates(session_dir: Path) -> Path:
    """Create an Excel file with date values."""
    xlsx_file = session_dir / "dates.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["name", "created_at"])
    ws.append(["Alice", datetime(2024, 1, 15, 10, 30, 0)])
    ws.append(["Bob", datetime(2024, 6, 20, 14, 45, 0)])
    wb.save(xlsx_file)
    wb.close()
    return xlsx_file


class TestExcelRead:
    """Tests for excel_read function."""

    def test_read_basic_xlsx(self, excel_read_fn, basic_xlsx, tmp_path):
        """Read a basic Excel file successfully."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert result["success"] is True
        assert result["columns"] == ["name", "age", "city"]
        assert result["column_count"] == 3
        assert result["row_count"] == 3
        assert result["total_rows"] == 3
        assert len(result["rows"]) == 3
        assert result["rows"][0] == {"name": "Alice", "age": 30, "city": "NYC"}
        assert result["sheet_name"] == "Sheet1"

    def test_read_specific_sheet(self, excel_read_fn, multi_sheet_xlsx, tmp_path):
        """Read a specific sheet from an Excel file."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="multi_sheet.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                sheet="Products",
            )

        assert result["success"] is True
        assert result["sheet_name"] == "Products"
        assert result["columns"] == ["id", "name", "price"]
        assert result["row_count"] == 2
        assert result["rows"][0]["name"] == "Widget"

    def test_read_nonexistent_sheet_error(self, excel_read_fn, multi_sheet_xlsx, tmp_path):
        """Return error for non-existent sheet."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="multi_sheet.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                sheet="NonExistent",
            )

        assert "error" in result
        assert "not found" in result["error"].lower()
        assert "Available sheets" in result["error"]

    def test_read_with_limit(self, excel_read_fn, basic_xlsx, tmp_path):
        """Read Excel with row limit."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                limit=2,
            )

        assert result["success"] is True
        assert result["row_count"] == 2
        assert result["total_rows"] == 3
        assert result["limit"] == 2
        assert len(result["rows"]) == 2
        assert result["rows"][0]["name"] == "Alice"
        assert result["rows"][1]["name"] == "Bob"

    def test_read_with_offset(self, excel_read_fn, basic_xlsx, tmp_path):
        """Read Excel with row offset."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                offset=1,
            )

        assert result["success"] is True
        assert result["row_count"] == 2
        assert result["offset"] == 1
        assert result["rows"][0]["name"] == "Bob"
        assert result["rows"][1]["name"] == "Charlie"

    def test_read_with_limit_and_offset(self, excel_read_fn, large_xlsx, tmp_path):
        """Read Excel with both limit and offset (pagination)."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="large.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                limit=10,
                offset=50,
            )

        assert result["success"] is True
        assert result["row_count"] == 10
        assert result["total_rows"] == 100
        assert result["offset"] == 50
        assert result["limit"] == 10
        # First row should be id=50
        assert result["rows"][0] == {"id": 50, "value": 500}

    def test_file_not_found(self, excel_read_fn, session_dir, tmp_path):
        """Return error for non-existent file."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="nonexistent.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result
        assert "not found" in result["error"].lower()

    def test_non_xlsx_extension(self, excel_read_fn, session_dir, tmp_path):
        """Return error for non-Excel file extension."""
        # Create a text file
        txt_file = session_dir / "data.txt"
        txt_file.write_text("name,age\nAlice,30\n", encoding="utf-8")

        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="data.txt",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result
        assert ".xlsx" in result["error"].lower() or ".xlsm" in result["error"].lower()

    def test_empty_xlsx_file(self, excel_read_fn, empty_xlsx, tmp_path):
        """Read empty Excel file (returns empty result)."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="empty.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert result["success"] is True
        assert result["row_count"] == 0
        assert result["rows"] == []

    def test_headers_only_xlsx(self, excel_read_fn, headers_only_xlsx, tmp_path):
        """Read Excel with only headers (no data rows)."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="headers_only.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert result["success"] is True
        assert result["columns"] == ["name", "age", "city"]
        assert result["row_count"] == 0
        assert result["total_rows"] == 0
        assert result["rows"] == []

    def test_missing_workspace_id(self, excel_read_fn, basic_xlsx, tmp_path):
        """Return error when workspace_id is missing."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="basic.xlsx",
                workspace_id="",
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result

    def test_missing_agent_id(self, excel_read_fn, basic_xlsx, tmp_path):
        """Return error when agent_id is missing."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id="",
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result

    def test_missing_session_id(self, excel_read_fn, basic_xlsx, tmp_path):
        """Return error when session_id is missing."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id="",
            )

        assert "error" in result

    def test_path_traversal_blocked(self, excel_read_fn, session_dir, tmp_path):
        """Prevent path traversal attacks."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="../../../etc/passwd.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result

    def test_negative_limit(self, excel_read_fn, basic_xlsx, tmp_path):
        """Return error for negative limit."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                limit=-1,
            )

        assert "error" in result
        assert "non-negative" in result["error"].lower()

    def test_negative_offset(self, excel_read_fn, basic_xlsx, tmp_path):
        """Return error for negative offset."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                offset=-1,
            )

        assert "error" in result
        assert "non-negative" in result["error"].lower()

    def test_offset_beyond_rows(self, excel_read_fn, basic_xlsx, tmp_path):
        """Offset beyond available rows returns empty result."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                offset=100,
            )

        assert result["success"] is True
        assert result["row_count"] == 0
        assert result["rows"] == []
        assert result["total_rows"] == 3

    def test_read_with_dates(self, excel_read_fn, xlsx_with_dates, tmp_path):
        """Read Excel with date values (should serialize to ISO format)."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_read_fn(
                path="dates.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert result["success"] is True
        # Dates should be serialized as ISO strings
        assert "2024-01-15" in result["rows"][0]["created_at"]


class TestExcelWrite:
    """Tests for excel_write function."""

    def test_write_new_xlsx(self, excel_tools, session_dir, tmp_path):
        """Write a new Excel file successfully."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_write"](
                path="output.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["name", "age", "city"],
                rows=[
                    {"name": "Alice", "age": 30, "city": "NYC"},
                    {"name": "Bob", "age": 25, "city": "LA"},
                ],
            )

        assert result["success"] is True
        assert result["columns"] == ["name", "age", "city"]
        assert result["column_count"] == 3
        assert result["rows_written"] == 2
        assert result["sheet_name"] == "Sheet1"

        # Verify file exists
        assert (session_dir / "output.xlsx").exists()

    def test_write_with_custom_sheet_name(self, excel_tools, session_dir, tmp_path):
        """Write Excel with custom sheet name."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_write"](
                path="output.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["id", "value"],
                rows=[{"id": 1, "value": 100}],
                sheet="MyData",
            )

        assert result["success"] is True
        assert result["sheet_name"] == "MyData"

    def test_write_creates_parent_directories(self, excel_tools, session_dir, tmp_path):
        """Write creates parent directories if needed."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_write"](
                path="subdir/nested/output.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["id"],
                rows=[{"id": 1}],
            )

        assert result["success"] is True
        assert (session_dir / "subdir" / "nested" / "output.xlsx").exists()

    def test_write_empty_columns_error(self, excel_tools, session_dir, tmp_path):
        """Return error when columns is empty."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_write"](
                path="output.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=[],
                rows=[],
            )

        assert "error" in result
        assert "empty" in result["error"].lower()

    def test_write_non_xlsx_extension_error(self, excel_tools, session_dir, tmp_path):
        """Return error for non-Excel file extension."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_write"](
                path="output.txt",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["id"],
                rows=[],
            )

        assert "error" in result
        assert ".xlsx" in result["error"].lower() or ".xlsm" in result["error"].lower()

    def test_write_empty_rows(self, excel_tools, session_dir, tmp_path):
        """Write Excel with headers but no rows."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_write"](
                path="output.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["name", "age"],
                rows=[],
            )

        assert result["success"] is True
        assert result["rows_written"] == 0


class TestExcelAppend:
    """Tests for excel_append function."""

    def test_append_to_existing_xlsx(self, excel_tools, basic_xlsx, tmp_path):
        """Append rows to an existing Excel file."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_append"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                rows=[
                    {"name": "David", "age": 28, "city": "Seattle"},
                    {"name": "Eve", "age": 32, "city": "Boston"},
                ],
            )

        assert result["success"] is True
        assert result["rows_appended"] == 2
        assert result["total_rows"] == 5

    def test_append_to_specific_sheet(self, excel_tools, multi_sheet_xlsx, tmp_path):
        """Append rows to a specific sheet."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_append"](
                path="multi_sheet.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                rows=[{"id": 3, "name": "Doohickey", "price": 49.99}],
                sheet="Products",
            )

        assert result["success"] is True
        assert result["sheet_name"] == "Products"
        assert result["rows_appended"] == 1

    def test_append_file_not_found(self, excel_tools, session_dir, tmp_path):
        """Return error when file doesn't exist."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_append"](
                path="nonexistent.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                rows=[{"name": "Alice"}],
            )

        assert "error" in result
        assert "not found" in result["error"].lower()

    def test_append_empty_rows_error(self, excel_tools, basic_xlsx, tmp_path):
        """Return error when rows is empty."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_append"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                rows=[],
            )

        assert "error" in result
        assert "empty" in result["error"].lower()

    def test_append_non_xlsx_extension_error(self, excel_tools, session_dir, tmp_path):
        """Return error for non-Excel file extension."""
        txt_file = session_dir / "data.txt"
        txt_file.write_text("name\nAlice\n", encoding="utf-8")

        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_append"](
                path="data.txt",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                rows=[{"name": "Bob"}],
            )

        assert "error" in result
        assert ".xlsx" in result["error"].lower() or ".xlsm" in result["error"].lower()


class TestExcelInfo:
    """Tests for excel_info function."""

    def test_get_info_basic_xlsx(self, excel_tools, basic_xlsx, tmp_path):
        """Get info about a basic Excel file."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_info"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert result["success"] is True
        assert result["sheet_count"] == 1
        assert result["sheet_names"] == ["Sheet1"]
        assert "file_size_bytes" in result
        assert result["file_size_bytes"] > 0
        assert len(result["sheets"]) == 1
        assert result["sheets"][0]["name"] == "Sheet1"
        assert result["sheets"][0]["columns"] == ["name", "age", "city"]
        assert result["sheets"][0]["row_count"] == 3

    def test_get_info_multi_sheet_xlsx(self, excel_tools, multi_sheet_xlsx, tmp_path):
        """Get info about a multi-sheet Excel file."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_info"](
                path="multi_sheet.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert result["success"] is True
        assert result["sheet_count"] == 3
        assert "Employees" in result["sheet_names"]
        assert "Products" in result["sheet_names"]
        assert "Summary" in result["sheet_names"]

    def test_get_info_file_not_found(self, excel_tools, session_dir, tmp_path):
        """Return error when file doesn't exist."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_info"](
                path="nonexistent.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result
        assert "not found" in result["error"].lower()

    def test_get_info_non_xlsx_extension_error(self, excel_tools, session_dir, tmp_path):
        """Return error for non-Excel file extension."""
        txt_file = session_dir / "data.txt"
        txt_file.write_text("name\nAlice\n", encoding="utf-8")

        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_info"](
                path="data.txt",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result
        assert ".xlsx" in result["error"].lower() or ".xlsm" in result["error"].lower()


class TestExcelSheetList:
    """Tests for excel_sheet_list function."""

    def test_list_sheets_basic(self, excel_tools, basic_xlsx, tmp_path):
        """List sheets in a basic Excel file."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_sheet_list"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert result["success"] is True
        assert result["sheet_names"] == ["Sheet1"]
        assert result["sheet_count"] == 1

    def test_list_sheets_multi_sheet(self, excel_tools, multi_sheet_xlsx, tmp_path):
        """List sheets in a multi-sheet Excel file."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_sheet_list"](
                path="multi_sheet.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert result["success"] is True
        assert result["sheet_count"] == 3
        assert "Employees" in result["sheet_names"]
        assert "Products" in result["sheet_names"]
        assert "Summary" in result["sheet_names"]

    def test_list_sheets_file_not_found(self, excel_tools, session_dir, tmp_path):
        """Return error when file doesn't exist."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_sheet_list"](
                path="nonexistent.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result
        assert "not found" in result["error"].lower()

    def test_list_sheets_non_xlsx_extension_error(self, excel_tools, session_dir, tmp_path):
        """Return error for non-Excel file extension."""
        txt_file = session_dir / "data.txt"
        txt_file.write_text("name\nAlice\n", encoding="utf-8")

        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_sheet_list"](
                path="data.txt",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert "error" in result
        assert ".xlsx" in result["error"].lower() or ".xlsm" in result["error"].lower()


class TestExcelIntegration:
    """Integration tests for Excel tools (write + read)."""

    def test_write_then_read(self, excel_tools, session_dir, tmp_path):
        """Write and then read back the same data."""
        test_data = [
            {"name": "Alice", "score": 95},
            {"name": "Bob", "score": 87},
            {"name": "Charlie", "score": 92},
        ]

        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            # Write
            write_result = excel_tools["excel_write"](
                path="test.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["name", "score"],
                rows=test_data,
            )
            assert write_result["success"] is True

            # Read back
            read_result = excel_tools["excel_read"](
                path="test.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert read_result["success"] is True
        assert read_result["row_count"] == 3
        assert read_result["rows"][0]["name"] == "Alice"
        assert read_result["rows"][0]["score"] == 95

    def test_write_append_read(self, excel_tools, session_dir, tmp_path):
        """Write, append, and then read back all data."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            # Write initial data
            excel_tools["excel_write"](
                path="test.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                columns=["id", "value"],
                rows=[{"id": 1, "value": "A"}, {"id": 2, "value": "B"}],
            )

            # Append more data
            excel_tools["excel_append"](
                path="test.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                rows=[{"id": 3, "value": "C"}, {"id": 4, "value": "D"}],
            )

            # Read back
            read_result = excel_tools["excel_read"](
                path="test.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
            )

        assert read_result["success"] is True
        assert read_result["row_count"] == 4
        assert read_result["rows"][2]["id"] == 3
        assert read_result["rows"][3]["value"] == "D"


# Check if duckdb is available for SQL tests
duckdb_available = importlib.util.find_spec("duckdb") is not None


@pytest.mark.skipif(not duckdb_available, reason="duckdb not installed")
class TestExcelSql:
    """Tests for excel_sql function."""

    def test_sql_basic_query(self, excel_tools, basic_xlsx, tmp_path):
        """Run basic SQL query on Excel file."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_sql"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                query="SELECT * FROM data",
            )

        assert result["success"] is True
        assert result["row_count"] == 3
        assert "name" in result["columns"]

    def test_sql_with_filter(self, excel_tools, basic_xlsx, tmp_path):
        """Run SQL query with WHERE clause."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_sql"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                query="SELECT * FROM data WHERE age > 25",
            )

        assert result["success"] is True
        assert result["row_count"] == 2  # Alice (30) and Charlie (35)

    def test_sql_with_aggregation(self, excel_tools, basic_xlsx, tmp_path):
        """Run SQL query with aggregation."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_sql"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                query="SELECT COUNT(*) as count, AVG(age) as avg_age FROM data",
            )

        assert result["success"] is True
        assert result["row_count"] == 1
        assert result["rows"][0]["count"] == 3

    def test_sql_specific_sheet(self, excel_tools, multi_sheet_xlsx, tmp_path):
        """Run SQL query on specific sheet."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_sql"](
                path="multi_sheet.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                query="SELECT * FROM data WHERE price > 100",
                sheet="Products",
            )

        assert result["success"] is True
        assert result["target_sheet"] == "Products"
        assert result["row_count"] == 1  # Gadget at 149.99

    def test_sql_join_sheets(self, excel_tools, multi_sheet_xlsx, tmp_path):
        """Join data across multiple sheets."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_sql"](
                path="multi_sheet.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                query="SELECT e.name, p.name as product FROM Employees e, Products p",
            )

        assert result["success"] is True
        # Cross join: 2 employees x 2 products = 4 rows
        assert result["row_count"] == 4

    def test_sql_empty_query_error(self, excel_tools, basic_xlsx, tmp_path):
        """Return error for empty query."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_sql"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                query="",
            )

        assert "error" in result
        assert "empty" in result["error"].lower()

    def test_sql_non_select_rejected(self, excel_tools, basic_xlsx, tmp_path):
        """Reject non-SELECT queries for security."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_sql"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                query="DELETE FROM data",
            )

        assert "error" in result
        assert "SELECT" in result["error"]

    def test_sql_drop_blocked(self, excel_tools, basic_xlsx, tmp_path):
        """Block DROP statements."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_sql"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                query="DROP TABLE data",
            )

        assert "error" in result

    def test_sql_insert_blocked(self, excel_tools, basic_xlsx, tmp_path):
        """Block INSERT statements."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_sql"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                query="INSERT INTO data VALUES ('x', 1, 'y')",
            )

        assert "error" in result

    def test_sql_file_not_found(self, excel_tools, session_dir, tmp_path):
        """Return error when file doesn't exist."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_sql"](
                path="nonexistent.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                query="SELECT * FROM data",
            )

        assert "error" in result
        assert "not found" in result["error"].lower()


class TestExcelSearch:
    """Tests for excel_search function."""

    def test_search_basic_contains(self, excel_tools, basic_xlsx, tmp_path):
        """Search for text containing a term."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_search"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                search_term="Alice",
            )

        assert result["success"] is True
        assert result["match_count"] >= 1
        assert any(m["value"] == "Alice" for m in result["matches"])

    def test_search_case_insensitive(self, excel_tools, basic_xlsx, tmp_path):
        """Search is case-insensitive by default."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_search"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                search_term="alice",
                case_sensitive=False,
            )

        assert result["success"] is True
        assert result["match_count"] >= 1

    def test_search_case_sensitive(self, excel_tools, basic_xlsx, tmp_path):
        """Case-sensitive search."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_search"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                search_term="alice",
                case_sensitive=True,
            )

        # "alice" (lowercase) won't match "Alice"
        assert result["success"] is True
        assert result["match_count"] == 0

    def test_search_exact_match(self, excel_tools, basic_xlsx, tmp_path):
        """Search with exact match."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_search"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                search_term="NYC",
                match_type="exact",
            )

        assert result["success"] is True
        assert result["match_count"] == 1
        assert result["matches"][0]["value"] == "NYC"

    def test_search_starts_with(self, excel_tools, basic_xlsx, tmp_path):
        """Search with starts_with match."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_search"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                search_term="Ch",
                match_type="starts_with",
            )

        assert result["success"] is True
        # Should match "Charlie" and "Chicago"
        assert result["match_count"] == 2

    def test_search_across_sheets(self, excel_tools, multi_sheet_xlsx, tmp_path):
        """Search across all sheets."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_search"](
                path="multi_sheet.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                search_term="Alice",
            )

        assert result["success"] is True
        assert result["match_count"] >= 1
        # Should search all sheets
        assert len(result["sheets_searched"]) == 3

    def test_search_specific_sheet(self, excel_tools, multi_sheet_xlsx, tmp_path):
        """Search in specific sheet only."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_search"](
                path="multi_sheet.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                search_term="Widget",
                sheet="Products",
            )

        assert result["success"] is True
        assert result["sheets_searched"] == ["Products"]
        assert result["match_count"] >= 1

    def test_search_skips_header_row(self, excel_tools, basic_xlsx, tmp_path):
        """Search should not match column header names."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_search"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                search_term="name",
                match_type="exact",
            )

        assert result["success"] is True
        assert result["match_count"] == 0

    def test_search_no_matches(self, excel_tools, basic_xlsx, tmp_path):
        """Search returns empty when no matches."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_search"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                search_term="ZZZZNOTFOUND",
            )

        assert result["success"] is True
        assert result["match_count"] == 0
        assert result["matches"] == []

    def test_search_empty_term_error(self, excel_tools, basic_xlsx, tmp_path):
        """Return error for empty search term."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_search"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                search_term="",
            )

        assert "error" in result
        assert "empty" in result["error"].lower()

    def test_search_invalid_match_type(self, excel_tools, basic_xlsx, tmp_path):
        """Return error for invalid match_type."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_search"](
                path="basic.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                search_term="test",
                match_type="invalid",
            )

        assert "error" in result
        assert "match_type" in result["error"]

    def test_search_file_not_found(self, excel_tools, session_dir, tmp_path):
        """Return error when file doesn't exist."""
        with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
            result = excel_tools["excel_search"](
                path="nonexistent.xlsx",
                workspace_id=TEST_WORKSPACE_ID,
                agent_id=TEST_AGENT_ID,
                session_id=TEST_SESSION_ID,
                search_term="test",
            )

        assert "error" in result
        assert "not found" in result["error"].lower()
